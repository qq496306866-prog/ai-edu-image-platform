from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook


TITLE_COLUMNS = {"商品标题", "标题", "title", "name"}
PROMPT_COLUMNS = {"提示词", "prompt"}
REFERENCE_IMAGE_COLUMNS = {"参考图片路径", "参考图", "image_path", "reference_image"}
SPREADSHEET_NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


@dataclass(frozen=True)
class ParsedExcelItem:
    row_number: int
    title: str
    prompt: str
    reference_image_path: str | None


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _find_column(headers: list[str], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if header in normalized_candidates:
            return index
    return None


def _read_shared_strings(archive: ZipFile) -> list[str]:
    try:
        xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ElementTree.fromstring(xml)
    strings: list[str] = []
    for item in root.findall("s:si", SPREADSHEET_NS):
        parts = [node.text or "" for node in item.findall(".//s:t", SPREADSHEET_NS)]
        strings.append("".join(parts))
    return strings


def _cell_text(cell: ElementTree.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value = cell.find("s:v", SPREADSHEET_NS)
        if value is None or value.text is None:
            return ""
        index = int(value.text)
        return shared_strings[index] if index < len(shared_strings) else ""
    if cell_type == "inlineStr":
        parts = [node.text or "" for node in cell.findall(".//s:t", SPREADSHEET_NS)]
        return "".join(parts)

    value = cell.find("s:v", SPREADSHEET_NS)
    return value.text if value is not None and value.text is not None else ""


def _column_index(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha())
    index = 0
    for letter in letters.upper():
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return max(index - 1, 0)


def _read_first_sheet_rows(path: Path) -> list[list[str]]:
    with ZipFile(path) as archive:
        shared_strings = _read_shared_strings(archive)
        xml = archive.read("xl/worksheets/sheet1.xml")

    root = ElementTree.fromstring(xml)
    rows: list[list[str]] = []
    for row in root.findall(".//s:sheetData/s:row", SPREADSHEET_NS):
        values: list[str] = []
        for cell in row.findall("s:c", SPREADSHEET_NS):
            cell_ref = cell.attrib.get("r", "")
            index = _column_index(cell_ref)
            while len(values) <= index:
                values.append("")
            values[index] = _cell_text(cell, shared_strings).strip()
        rows.append(values)
    return rows


def _parse_rows(rows: list[list[object]]) -> list[ParsedExcelItem]:
    if not rows:
        return []

    header_row = rows[0]
    headers = [_normalize_header(cell) for cell in header_row]
    title_index = _find_column(headers, TITLE_COLUMNS)
    prompt_index = _find_column(headers, PROMPT_COLUMNS)
    reference_index = _find_column(headers, REFERENCE_IMAGE_COLUMNS)

    if title_index is None or prompt_index is None:
        raise ValueError("Excel must include title and prompt columns")

    parsed_items: list[ParsedExcelItem] = []
    for row_number, row in enumerate(rows[1:], start=2):
        title = str(row[title_index] or "").strip() if title_index < len(row) else ""
        prompt = str(row[prompt_index] or "").strip() if prompt_index < len(row) else ""
        reference_image_path = (
            str(row[reference_index] or "").strip()
            if reference_index is not None and reference_index < len(row)
            else ""
        )

        if not title and not prompt and not reference_image_path:
            continue
        if not title or not prompt:
            continue

        parsed_items.append(
            ParsedExcelItem(
                row_number=row_number,
                title=title,
                prompt=prompt,
                reference_image_path=reference_image_path or None,
            )
        )

    return parsed_items


def parse_generation_excel(path: Path) -> list[ParsedExcelItem]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]

    try:
        return _parse_rows(rows)
    except ValueError:
        return _parse_rows(_read_first_sheet_rows(path))
